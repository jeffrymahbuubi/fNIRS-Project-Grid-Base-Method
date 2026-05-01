#!/usr/bin/env python3
"""
extract_metrics.py — Extract performance metrics from ViT experiment .pkl files.

Usage:
    python scripts/extract_metrics.py [TARGET_DIR] [--xlsx] [--output OUTPUT]

Arguments:
    TARGET_DIR   Root directory containing 5fold/, 10fold/, loso/ subdirs.
                 Defaults to: research/experiments/20260428

Options:
    --xlsx       Save as .xlsx with one sheet per CV strategy (default: .csv per strategy).
    --output     Output file base name (without extension). Default: experiment_metrics.
"""

import argparse
import glob
import math
import os
import pickle
import re
import sys
from pathlib import Path

import numpy as np

# ── metric computation ────────────────────────────────────────────────────────

def metrics_from_cm(cm):
    """Compute all metrics from a 2×2 confusion matrix [[TN,FP],[FN,TP]]."""
    TN, FP = int(cm[0, 0]), int(cm[0, 1])
    FN, TP = int(cm[1, 0]), int(cm[1, 1])

    acc   = (TP + TN) / (TP + TN + FP + FN) if (TP + TN + FP + FN) > 0 else float("nan")
    sens  = TP / (TP + FN)                   if (TP + FN) > 0            else float("nan")
    spec  = TN / (TN + FP)                   if (TN + FP) > 0            else float("nan")
    prec  = TP / (TP + FP)                   if (TP + FP) > 0            else float("nan")
    npv   = TN / (TN + FN)                   if (TN + FN) > 0            else float("nan")
    f1    = (2 * TP) / (2 * TP + FP + FN)   if (2 * TP + FP + FN) > 0  else float("nan")
    ba    = (sens + spec) / 2 if not (math.isnan(sens) or math.isnan(spec)) else float("nan")

    # Cohen's Kappa
    total = TP + TN + FP + FN
    if total > 0:
        po = (TP + TN) / total
        pe = ((TP + FP) * (TP + FN) + (TN + FN) * (TN + FP)) / (total ** 2)
        kappa = (po - pe) / (1 - pe) if (1 - pe) != 0 else float("nan")
    else:
        kappa = float("nan")

    # Matthews Correlation Coefficient (denom=0 → MCC=0 by sklearn convention)
    num   = TP * TN - FP * FN
    denom = math.sqrt((TP + FP) * (TP + FN) * (TN + FP) * (TN + FN))
    mcc   = num / denom if denom != 0 else 0.0

    return dict(acc=acc, sens=sens, spec=spec, prec=prec, npv=npv,
                f1=f1, ba=ba, kappa=kappa, mcc=mcc)


def fmt(mean, sd):
    if math.isnan(mean) or math.isnan(sd):
        return "—"
    return f"{mean:.3f}±{sd:.3f}"


def fmt_ci(mean, sd, n):
    if math.isnan(mean) or math.isnan(sd) or n == 0:
        return "—"
    half = 1.96 * sd / math.sqrt(n)
    return f"[{mean - half:.3f}, {mean + half:.3f}]"


# ── folder / file helpers ─────────────────────────────────────────────────────

FOLD_PKL_RE   = re.compile(r"ViT_fold_(\d+)\.pkl$")
SUBJ_PKL_RE   = re.compile(r"ViT_subject_(\w+)\.pkl$")
OVERALL_PKL   = re.compile(r"ViT_(kfold_overall|loso_overall)\.pkl$")
EXP_DIR_RE    = re.compile(
    r"ViT_(?P<task>\w+?)_(?P<signal>hbo|hbr|hbt)_"
    r"(?:kfold(?P<k>\d+)|loso)_",
    re.IGNORECASE,
)


def parse_exp_dir_name(dirname):
    """Return (task, signal, cv_label) from experiment directory name."""
    m = EXP_DIR_RE.search(os.path.basename(dirname))
    if not m:
        return None, None, None
    task   = m.group("task")
    signal = m.group("signal").upper()
    k      = m.group("k")
    cv     = f"{k}-fold" if k else "LOSO"
    return task, signal, cv


def load_fold_pkls(exp_dir):
    """Return list of per-fold metric dicts, sorted by fold number."""
    pkls = sorted(
        glob.glob(os.path.join(exp_dir, "*.pkl")),
        key=lambda p: int(m.group(1)) if (m := FOLD_PKL_RE.search(os.path.basename(p))) else -1,
    )
    fold_metrics = []
    for pkl_path in pkls:
        fname = os.path.basename(pkl_path)
        if not FOLD_PKL_RE.search(fname):
            continue
        with open(pkl_path, "rb") as f:
            d = pickle.load(f)
        cm = d.get("conf_matrix") if "conf_matrix" in d else d.get("confusion_matrix")
        if cm is None:
            continue
        m_dict = metrics_from_cm(np.array(cm))
        m_dict["best_epoch"] = d.get("best_epoch", float("nan"))
        fold_metrics.append(m_dict)
    return fold_metrics


def load_subject_pkls(exp_dir):
    """Return list of per-subject metric dicts for LOSO."""
    pkls = sorted(glob.glob(os.path.join(exp_dir, "ViT_subject_*.pkl")))
    subj_metrics = []
    for pkl_path in pkls:
        with open(pkl_path, "rb") as f:
            d = pickle.load(f)
        cm = d.get("conf_matrix") if "conf_matrix" in d else d.get("confusion_matrix")
        if cm is None:
            continue
        m_dict = metrics_from_cm(np.array(cm))
        subj_metrics.append(m_dict)
    return subj_metrics


def load_overall_pkl(exp_dir):
    """Return (overall_acc, overall_f1, cm_metrics) from pooled CM pkl."""
    for fname in os.listdir(exp_dir):
        if OVERALL_PKL.search(fname):
            with open(os.path.join(exp_dir, fname), "rb") as f:
                d = pickle.load(f)
            cm = d.get("confusion_matrix")
            if cm is None:
                return float("nan"), float("nan"), {}
            m = metrics_from_cm(np.array(cm))
            return m["acc"], m["f1"], m
    return float("nan"), float("nan"), {}


# ── aggregation ───────────────────────────────────────────────────────────────

METRIC_KEYS = ["acc", "sens", "spec", "prec", "npv", "f1", "ba", "kappa", "mcc"]
METRIC_LABELS = {
    "acc":   "Acc",
    "sens":  "Sens",
    "spec":  "Spec",
    "prec":  "Prec",
    "npv":   "NPV",
    "f1":    "F1",
    "ba":    "BA",
    "kappa": "κ",
    "mcc":   "MCC",
}


def aggregate(fold_list):
    """Return {metric: (mean, sd)} ignoring NaN values."""
    result = {}
    for key in METRIC_KEYS:
        vals = [f[key] for f in fold_list if not math.isnan(f.get(key, float("nan")))]
        if vals:
            result[key] = (float(np.mean(vals)), float(np.std(vals, ddof=1)))
        else:
            result[key] = (float("nan"), float("nan"))
    return result


def build_row(task, signal, cv, agg, n_folds, overall_acc, overall_f1, overall_m):
    """Build a flat dict row for the summary table."""
    row = {"Task": task, "Signal": signal, "CV": cv}
    for key in METRIC_KEYS:
        mean, sd = agg[key]
        label = METRIC_LABELS[key]
        row[f"{label} Mean"] = round(mean, 4) if not math.isnan(mean) else None
        row[f"{label} SD"]   = round(sd,   4) if not math.isnan(sd)   else None
        row[f"{label} Mean±SD"] = fmt(mean, sd)
        row[f"{label} 95% CI"]  = fmt_ci(mean, sd, n_folds)
    row["Overall Acc"]       = round(float(overall_acc), 4) if not math.isnan(overall_acc) else None
    row["Overall F1"]        = round(float(overall_f1),  4) if not math.isnan(overall_f1)  else None
    row["Overall Sens"]      = round(overall_m.get("sens", float("nan")), 4) if overall_m else None
    row["Overall Spec"]      = round(overall_m.get("spec", float("nan")), 4) if overall_m else None
    row["Overall Prec"]      = round(overall_m.get("prec", float("nan")), 4) if overall_m else None
    row["Overall NPV"]       = round(overall_m.get("npv",  float("nan")), 4) if overall_m else None
    row["Overall BA"]        = round(overall_m.get("ba",   float("nan")), 4) if overall_m else None
    row["Overall κ"]         = round(overall_m.get("kappa",float("nan")), 4) if overall_m else None
    row["Overall MCC"]       = round(overall_m.get("mcc",  float("nan")), 4) if overall_m else None
    row["N Folds/Subjects"]  = n_folds
    return row


# ── CV-specific processing ────────────────────────────────────────────────────

def process_kfold_dir(cv_dir):
    """Process all experiment subdirectories in a k-fold CV directory."""
    rows = []
    for exp_dir in sorted(glob.glob(os.path.join(cv_dir, "ViT_*"))):
        if not os.path.isdir(exp_dir):
            continue
        task, signal, cv = parse_exp_dir_name(exp_dir)
        if task is None:
            continue
        fold_list = load_fold_pkls(exp_dir)
        if not fold_list:
            print(f"  [WARN] No fold pkls found in {exp_dir}", file=sys.stderr)
            continue
        agg = aggregate(fold_list)
        overall_acc, overall_f1, overall_m = load_overall_pkl(exp_dir)
        row = build_row(task, signal, cv, agg, len(fold_list), overall_acc, overall_f1, overall_m)
        rows.append(row)
        print(f"  [{cv}] {task} {signal}: {len(fold_list)} folds, "
              f"Acc={agg['acc'][0]:.3f}±{agg['acc'][1]:.3f}, "
              f"Overall Acc={overall_acc:.3f}")
    return rows


def process_loso_dir(loso_dir):
    """Process all LOSO experiment subdirectories."""
    rows = []
    for exp_dir in sorted(glob.glob(os.path.join(loso_dir, "ViT_*"))):
        if not os.path.isdir(exp_dir):
            continue
        task, signal, cv = parse_exp_dir_name(exp_dir)
        if task is None:
            continue
        subj_list = load_subject_pkls(exp_dir)
        if not subj_list:
            print(f"  [WARN] No subject pkls in {exp_dir}", file=sys.stderr)
            continue
        agg = aggregate(subj_list)
        overall_acc, overall_f1, overall_m = load_overall_pkl(exp_dir)
        row = build_row(task, signal, "LOSO", agg, len(subj_list), overall_acc, overall_f1, overall_m)
        rows.append(row)
        print(f"  [LOSO] {task} {signal}: {len(subj_list)} subjects, "
              f"Overall Acc={overall_acc:.3f}")
    return rows


# ── output helpers ────────────────────────────────────────────────────────────

# Columns in the final summary table (mean±SD and overall side by side)
SUMMARY_COLS = (
    ["Task", "Signal", "CV", "N Folds/Subjects"]
    + [f"{METRIC_LABELS[k]} Mean±SD" for k in METRIC_KEYS]
    + ["Overall Acc", "Overall F1",
       "Overall Sens", "Overall Spec", "Overall Prec", "Overall NPV",
       "Overall BA", "Overall κ", "Overall MCC"]
)

# Per-metric detail columns (Mean, SD, CI separately for downstream analysis)
DETAIL_COLS = (
    ["Task", "Signal", "CV", "N Folds/Subjects"]
    + [col
       for k in METRIC_KEYS
       for col in (f"{METRIC_LABELS[k]} Mean", f"{METRIC_LABELS[k]} SD",
                   f"{METRIC_LABELS[k]} Mean±SD", f"{METRIC_LABELS[k]} 95% CI")]
    + ["Overall Acc", "Overall F1",
       "Overall Sens", "Overall Spec", "Overall Prec", "Overall NPV",
       "Overall BA", "Overall κ", "Overall MCC"]
)


def rows_to_csv(rows, path, cols):
    import csv
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    print(f"  → {path}")


def rows_to_sheet(ws, rows, cols):
    """Write rows to an openpyxl worksheet with a styled header."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E6099")

    for col_idx, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            ws.cell(row=row_idx, column=col_idx, value=val if val is not None else "")

    # Auto-width (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Extract ViT experiment metrics from .pkl files.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "target_dir",
        nargs="?",
        default=None,
        help="Root directory containing 5fold/, 10fold/, loso/. "
             "Defaults to research/experiments/20260428 relative to script location.",
    )
    parser.add_argument(
        "--xlsx",
        action="store_true",
        help="Save output as a single .xlsx file with one sheet per CV strategy.",
    )
    parser.add_argument(
        "--output",
        default="experiment_metrics",
        help="Output file base name (no extension). Default: experiment_metrics.",
    )
    args = parser.parse_args()

    # Resolve target directory
    if args.target_dir:
        base = Path(args.target_dir).expanduser().resolve()
    else:
        script_dir = Path(__file__).parent.resolve()
        base = script_dir.parent / "research" / "experiments" / "20260428"

    if not base.is_dir():
        print(f"ERROR: target directory not found: {base}", file=sys.stderr)
        sys.exit(1)

    print(f"Target directory : {base}")
    print(f"Output format    : {'xlsx' if args.xlsx else 'csv'}")
    print(f"Output base name : {args.output}")
    print()

    all_sheets = {}  # sheet_name → list[row]

    for subdir_name, label in [("5fold", "5-Fold"), ("10fold", "10-Fold"), ("loso", "LOSO")]:
        subdir = base / subdir_name
        if not subdir.is_dir():
            print(f"[SKIP] {subdir} not found")
            continue
        print(f"Processing {label} ...")
        if subdir_name == "loso":
            rows = process_loso_dir(str(subdir))
        else:
            rows = process_kfold_dir(str(subdir))
        all_sheets[label] = rows
        print()

    if not all_sheets:
        print("No data found — check target directory structure.", file=sys.stderr)
        sys.exit(1)

    # Resolve output path (save alongside script or to cwd if run from elsewhere)
    out_base = Path(args.output)
    if not out_base.is_absolute():
        out_base = Path.cwd() / out_base

    if args.xlsx:
        try:
            import openpyxl
        except ImportError:
            print("ERROR: openpyxl is required for --xlsx. Install with: pip install openpyxl",
                  file=sys.stderr)
            sys.exit(1)

        out_path = out_base.with_suffix(".xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        for sheet_name, rows in all_sheets.items():
            if not rows:
                continue
            # Summary sheet (Mean±SD + Overall)
            ws_summary = wb.create_sheet(f"{sheet_name} Summary")
            rows_to_sheet(ws_summary, rows, SUMMARY_COLS)

            # Detail sheet (numeric Mean/SD + CI)
            ws_detail = wb.create_sheet(f"{sheet_name} Detail")
            rows_to_sheet(ws_detail, rows, DETAIL_COLS)

        wb.save(out_path)
        print(f"\nSaved: {out_path}")

    else:
        print("Saving CSV files ...")
        for sheet_name, rows in all_sheets.items():
            if not rows:
                continue
            safe_name = sheet_name.replace("-", "").replace(" ", "_").lower()
            # Summary CSV
            summary_path = out_base.parent / f"{out_base.name}_{safe_name}_summary.csv"
            rows_to_csv(rows, summary_path, SUMMARY_COLS)
            # Detail CSV
            detail_path = out_base.parent / f"{out_base.name}_{safe_name}_detail.csv"
            rows_to_csv(rows, detail_path, DETAIL_COLS)

    print("\nDone.")


if __name__ == "__main__":
    main()
